"""
JollyCast Bot 150問フルテスト
Q1-Q150 全問テスト — 現在のKNOWLEDGEによる通算評価
Excel: jollycast_bot_test_results_v2.xlsx に "Round7（150問フル）" タブを追加
"""
import os, time, sys, re, json
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from zendesk_loader import search_articles

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

base = Path(__file__).parent

for line in (base / ".env").read_text(encoding="utf-8-sig").splitlines():
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1)
        os.environ[k.strip()] = v.strip()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
if not ANTHROPIC_API_KEY:
    print("ERROR: ANTHROPIC_API_KEY not found"); sys.exit(1)

import anthropic
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

src = (base / "app.py").read_text(encoding="utf-8")
KNOWLEDGE = re.search(r'KNOWLEDGE = """(.+?)"""', src, re.DOTALL).group(1)
TROUBLE_FLOW = re.search(r'TROUBLE_FLOW = """(.+?)"""', src, re.DOTALL).group(1)

cache_path = base / "zendesk_cache.json"
articles = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else []
print(f"Zendesk記事: {len(articles)}件読み込み")

# カテゴリ分類
TYPE_B_CATEGORIES: set[str] = set()
GTN_CATEGORIES: set[str] = {"GTN・送金", "GTN・銀行", "GTN・携帯"}


def get_bot_response(question: str) -> str:
    prompt = f"""You are a support assistant for JollyCast (ジョリーキャスト), helping Filipino cast members in Japan.
CRITICAL RULES:
1. Answer ONLY from the KNOWLEDGE BASE and TROUBLE FLOW provided below.
2. Do NOT use general knowledge or invent procedures not written in these materials.
3. If not covered: for emergencies (safety, serious damage mid-service) call HQ 📞 050-3183-8835; for non-urgent issues direct to the inquiry form (Cast App → 問い合わせフォーム).
4. Always respond in English. Be concise. Number steps clearly.

KNOWLEDGE BASE:
{KNOWLEDGE}

TROUBLE FLOW:
{TROUBLE_FLOW}

QUESTION: {question}"""
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}]
    )
    return msg.content[0].text


def _eval_sources(question: str) -> str:
    """評価者ソース: KNOWLEDGE + TROUBLE_FLOW + Zendesk関連記事上位5件"""
    hits = search_articles(question, articles, top_k=5)
    zendesk_excerpt = "\n\n".join(
        f"=== {a['title']} ===\n{a['content'][:2000]}" for a in hits
    )
    return (
        f"=== KNOWLEDGE BASE ===\n{KNOWLEDGE}\n\n"
        f"=== TROUBLE FLOW ===\n{TROUBLE_FLOW}\n\n"
        f"=== ZENDESK (top 5 relevant articles) ===\n{zendesk_excerpt}"
    )


def evaluate_response(question: str, response: str, category: str, sources: str) -> tuple[str, str]:
    if category in GTN_CATEGORIES:
        type_instruction = (
            "This is a life-admin question (banking, housing, SIM, etc.).\n"
            "CORRECT answer: direct the cast to GTN via their app.\n"
            "Answering with specific procedures = WRONG. HQ escalation = WRONG."
        )
        source_section = ""
    elif category in TYPE_B_CATEGORIES:
        type_instruction = (
            "This is a TYPE B general knowledge question (navigation, common sense, etc.).\n"
            "Evaluate whether the answer is practically correct and helpful.\n"
            "HQ escalation is NOT the expected answer for these questions."
        )
        source_section = ""
    else:
        type_instruction = (
            "This is a TYPE A business procedure question.\n"
            "The REFERENCE SOURCES below are the bot's ground truth.\n"
            "Penalize: (1) answers that invent procedures not in the sources,\n"
            "(2) unnecessary HQ phone calls when the answer IS clearly in the sources."
        )
        source_section = f"\n=== REFERENCE SOURCES ===\n{sources}\n"

    prompt = f"""You are evaluating a support bot response for JollyCast cast members (CaSy, Japan).
CaSy support number 050-3183-8835 is correct and appropriate to include for emergencies.

{type_instruction}
{source_section}
◎ = Accurate, specific, immediately actionable
○ = Mostly correct, minor gaps
△ = Partially correct, could mislead
✕ = Wrong or unable to answer

QUESTION: {question}

BOT RESPONSE:
{response}

Reply in this exact format (2 lines only):
RATING: [◎/○/△/✕]
COMMENT: [1-2 sentences in Japanese]"""
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    )
    text = msg.content[0].text.strip()
    rating, comment = "", ""
    for line in text.splitlines():
        if line.startswith("RATING:"):
            rating = line.replace("RATING:", "").strip()
        elif line.startswith("COMMENT:"):
            comment = line.replace("COMMENT:", "").strip()
    return rating, comment


def style_header(ws, headers, color="2E4057"):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", start_color=color, end_color=color)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30


def apply_row_style(ws, row_idx, cat, rating, num_cols):
    category_colors = {
        "お客様不在": "DDEEFF", "物損・事故": "FFE4E4", "キャンセル": "FFF3CD",
        "道に迷う・住所": "E4FFE4", "鍵・入室トラブル": "F0E4FF", "スケジュール変更": "FFE4F0",
        "QR・バウチャー": "E4FFFF", "サービス後・移動": "FFFAE4", "安全・緊急事態": "FFE8E8",
        "アプリ操作・報告": "E8F4FF", "その他ルール・マナー": "F0FFE8",
        "不衛生・退出判断": "FFE4CC", "自治体バウチャー": "CCF0FF", "Recoru・勤怠管理": "E8E8FF",
        "キャンセル詳細": "FFF0E0", "体調不良・突発欠勤": "EEFFEE", "料理サービス・返金": "FFF0F8",
    }
    rating_colors = {"◎": "D4EDDA", "○": "D1ECF1", "△": "FFF3CD", "✕": "F8D7DA"}
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cat_color = category_colors.get(cat, "FFFFFF")
    r_color = rating_colors.get(rating, "FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        fill_color = r_color if col == 6 else cat_color
        cell.fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
        cell.font = Font(name="Arial", size=9)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
    if num_cols >= 6:
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center", vertical="top")
    ws.row_dimensions[row_idx].height = 80


def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def main():
    questions_file = base / "jollycast_bot_test_questions.xlsx"
    results_file = base / "jollycast_bot_test_results_v2.xlsx"

    wb_q = load_workbook(questions_file)
    ws_q = wb_q.active
    all_rows = []
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        num, cat, q_en = row[0], row[1], row[2]
        q_ja = row[3] if len(row) > 3 else ""
        if num and q_en:
            all_rows.append((int(num), cat, q_en, q_ja or ""))

    # Q1-Q150 全問
    all_rows.sort(key=lambda x: x[0])
    print(f"テスト対象: {len(all_rows)}問 (Q1-Q150)")

    wb = load_workbook(results_file)

    for sheet_name in ["Round7（150問フル）", "Round7カテゴリ別"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws7 = wb.create_sheet("Round7（150問フル）")
    headers = ["#", "カテゴリ", "質問（English）", "質問（日本語）", "ボットの回答", "AI評価", "AI評価コメント", "担当者評価", "担当者コメント"]
    style_header(ws7, headers)
    set_col_widths(ws7, [4, 20, 40, 30, 50, 8, 35, 10, 30])
    ws7.freeze_panes = "A2"

    total = len(all_rows)
    all_ratings = []

    for i, (num, cat, q_en, q_ja) in enumerate(all_rows, 1):
        label = f"[{i}/{total}] Q{num}"
        print(f"{label}: {q_en[:55]}...", flush=True)
        response = get_bot_response(q_en)
        time.sleep(0.3)
        sources = _eval_sources(q_en)
        rating, comment = evaluate_response(q_en, response, cat or "", sources)
        time.sleep(0.3)
        mark = {"◎": "◎", "○": "○", "△": "△", "✕": "✕"}.get(rating, rating)
        print(f"  -> {mark}", flush=True)
        all_ratings.append((cat, rating))
        ws7.append([num, cat, q_en, q_ja, response, rating, comment, "", ""])
        apply_row_style(ws7, ws7.max_row, cat or "", rating, len(headers))
        wb.save(results_file)

    # カテゴリ別サマリー
    ws_cat = wb.create_sheet("Round7カテゴリ別")
    cat_headers = ["カテゴリ", "◎", "○", "△", "✕", "合計", "良評価率(◎+○)"]
    style_header(ws_cat, cat_headers, "1A3A5C")
    set_col_widths(ws_cat, [22, 8, 8, 8, 8, 8, 16])

    cat_ratings = defaultdict(list)
    for cat, rating in all_ratings:
        cat_ratings[cat or "未分類"].append(rating)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    print(f"\n=== Round7 カテゴリ別サマリー ===")
    for cat in sorted(cat_ratings.keys()):
        cnt = Counter(cat_ratings[cat])
        total_cat = sum(cnt.values())
        good = cnt.get("◎", 0) + cnt.get("○", 0)
        good_rate = f"{round(good/total_cat*100, 1)}%" if total_cat else "0%"
        print(f"  {cat}: ◎{cnt.get('◎',0)} ○{cnt.get('○',0)} △{cnt.get('△',0)} ✕{cnt.get('✕',0)} -> {good_rate}")
        ws_cat.append([cat, cnt.get("◎", 0), cnt.get("○", 0), cnt.get("△", 0), cnt.get("✕", 0), total_cat, good_rate])
        row_idx = ws_cat.max_row
        for col in range(1, 8):
            cell = ws_cat.cell(row=row_idx, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws_cat.row_dimensions[row_idx].height = 22

    r7_count = Counter(r for _, r in all_ratings)
    total_all = sum(r7_count.values())
    good_all = r7_count.get("◎", 0) + r7_count.get("○", 0)
    good_rate_all = f"{round(good_all/total_all*100, 1)}%" if total_all else "0%"
    ws_cat.append(["【合計】", r7_count.get("◎", 0), r7_count.get("○", 0), r7_count.get("△", 0), r7_count.get("✕", 0), total_all, good_rate_all])
    row_idx = ws_cat.max_row
    for col in range(1, 8):
        cell = ws_cat.cell(row=row_idx, column=col)
        cell.fill = PatternFill("solid", start_color="DDDDDD", end_color="DDDDDD")
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws_cat.row_dimensions[row_idx].height = 25

    wb.save(results_file)

    print(f"\n=== 全体 {total_all}問 ===")
    print(f"◎:{r7_count.get('◎',0)} ○:{r7_count.get('○',0)} △:{r7_count.get('△',0)} ✕:{r7_count.get('✕',0)}")
    print(f"良評価率: {good_rate_all}")
    print(f"\nExcel保存完了: {results_file}")


if __name__ == "__main__":
    main()
