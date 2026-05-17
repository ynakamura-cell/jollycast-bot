"""
JollyCast Bot スポットテスト（2026-05-17 新アーキテクチャ対応）
Part A: 20問 既存KNOWLEDGEシナリオ検証
Part B:  8問 新規追加KNOWLEDGEセクション検証（仕上がり不満・引越し・住所相違・パッケージプラン等）
Part C:  3問 GTN案内検証

全パート共通: KNOWLEDGE + TROUBLE_FLOW のみ（Zendesk不使用）
Excel: jollycast_bot_test_results_v2.xlsx に以下のタブを追加
  - Spot-A（既存KNOWLEDGE20問）
  - Spot-B（新規KNOWLEDGE8問）
  - Spot-C（GTN3問）
"""
import os, time, sys, re, json
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from zendesk_loader import search_articles

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

base = Path(__file__).parent

for line in (base / ".env").read_text(encoding="utf-8-sig").splitlines():
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1)
        os.environ[k.strip()] = v.strip()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
if not ANTHROPIC_API_KEY:
    print("ERROR: ANTHROPIC_API_KEY not found"); sys.exit(1)

import anthropic
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

src = (base / "app.py").read_text(encoding="utf-8")
KNOWLEDGE    = re.search(r'KNOWLEDGE = """(.+?)"""',    src, re.DOTALL).group(1)
TROUBLE_FLOW = re.search(r'TROUBLE_FLOW = """(.+?)"""', src, re.DOTALL).group(1)

cache_path = base / "zendesk_cache.json"
if cache_path.exists():
    articles = json.loads(cache_path.read_text(encoding="utf-8"))
    print(f"Zendesk記事: {len(articles)}件読み込み")
else:
    articles = []
    print("WARNING: zendesk_cache.json なし")

ZENDESK_CONTENT = "\n\n".join(
    f"=== {a['title']} ===\n{a['content']}" for a in articles
)

# ── Part A: 20問 ──────────────────────────────────────────────
QUESTIONS_A = [
    ("ダブルブッキング",   "I arrived at the customer's home and a Japanese cast is already there. What should I do?"),
    ("ダブルブッキング",   "I arrived and another JollyCast cast is there. The customer says they only need 1 cast. What happens to the fee?"),
    ("ダブルブッキング",   "Double booking — both casts are JollyCast. The customer says they only made one booking and refuses to pay for two. What do I say?"),
    ("サービス後・移動",   "My last service of the day ended early. Do I need to go to the office?"),
    ("サービス後・移動",   "I have a 2-hour gap between services. What should I do during this time?"),
    ("サービス後・移動",   "My afternoon services were all cancelled. What should I do for the rest of the day?"),
    ("キャンセル詳細",     "The customer messaged me at 17:50, two days before the service, to cancel. Is this free cancellation?"),
    ("不衛生・退出判断",   "I arrived and there is a severe cockroach infestation. Can I leave without asking the customer to clean first?"),
    ("不衛生・退出判断",   "I feel physically ill because of the smell in the customer's home. What do I do?"),
    ("自治体バウチャー",   "Who fills in the back of the municipal voucher after service — me or the customer?"),
    ("自治体バウチャー",   "The customer in Taito-ku has no smartphone and cannot scan the QR code. What do I do?"),
    ("QR・バウチャー",     "I forgot to have the customer scan the QR code at the end of service. What happens?"),
    ("QR・バウチャー",     "The QR code is not showing on my app screen. What do I do?"),
    ("料理サービス・返金", "I arrived 40 minutes late to a cooking service. What are the rules?"),
    ("料理サービス・返金", "The customer says the food doesn't taste good and wants a full refund. What do I say?"),
    ("お客様不在",         "I've been waiting outside for 30 minutes and the customer is still not responding. What do I do now?"),
    ("物損・事故",         "I accidentally cracked a window — it affects the customer's daily life. What do I do first?"),
    ("HQエスカレーション", "There is a strong gas smell in the customer's home. What should I do?"),
    ("HQエスカレーション", "The customer asked me to teach them how to clean properly while I work. What do I say?"),
    ("HQエスカレーション", "I left my cleaning tools at the customer's home after service. How do I get them back?"),
]

# ── Part C: 3問（GTN案内が正解） ─────────────────────────────
QUESTIONS_C = [
    ("GTN・送金", "I want to send money to my family in the Philippines. How can I do this in Japan?"),
    ("GTN・銀行", "I don't have a Japanese bank account yet. What should I do?"),
    ("GTN・携帯", "My phone SIM card is not working. Who should I contact?"),
]

# ── Part B: 新KNOWLEDGE検証（2026-05-17追記セクション） ────────
QUESTIONS_B = [
    # 既存（KNOWLEDGEに追記済みを確認）
    ("自治体バウチャー",   "I'm doing a service in Musashino-shi (武蔵野市). The customer has municipal vouchers. What do I do with them?"),
    ("アプリ操作・報告",   "I made a mistake in my daily report after submitting it. Can I correct it, and how?"),
    ("鍵・入室トラブル",   "The customer has a key box (キーボックス) outside their home. How do I access it and return the key after service?"),
    # 新追記セクションの検証
    ("仕上がり不満",       "After I finished cleaning, the customer said they are not happy with my work. What should I do?"),
    ("引越しサポート",     "The customer asked me to help them pack boxes because they are moving next week. Can I do this, and are there any restrictions?"),
    ("住所相違",           "When I arrived, the customer asked me to go to a different address nearby — their friend's apartment. Can I do this?"),
    ("パッケージプラン",   "The customer is asking me to make the seasoning lighter because their child doesn't like strong flavors. What do I say?"),
    ("フリープラン混同",   "I heard there is a shopping proxy service (買物代行). Can I use this option for today's cooking service?"),
]

# ── カテゴリ分類 ─────────────────────────────────────────────
# TYPE B: 一般知識で回答（CaSyマニュアル外）
TYPE_B_CATEGORIES: set[str] = set()   # 例: {"電車・交通", "住所・建物ナビ"}

# GTN: 非急ぎの生活インフラ系 → GTN案内が正解
GTN_CATEGORIES: set[str] = {"GTN・送金", "GTN・銀行", "GTN・携帯"}


def api_call_with_retry(fn, max_retries=6):
    for attempt in range(max_retries):
        try:
            return fn()
        except anthropic.RateLimitError:
            wait = 65 * (attempt + 1)
            print(f"  [Rate limit] {wait}秒待機...", flush=True)
            time.sleep(wait)
    raise RuntimeError("リトライ上限に達しました")


def get_bot_response_a(question: str) -> str:
    """Part A: KNOWLEDGE + TROUBLE_FLOW のみ"""
    prompt = f"""You are a support assistant for JollyCast (ジョリーキャスト), helping Filipino cast members in Japan.

CRITICAL RULES:
1. Answer ONLY from the KNOWLEDGE BASE and TROUBLE FLOW provided below.
2. Do NOT use general knowledge or invent procedures not written in these materials.
3. If not covered: for emergencies (safety, serious damage mid-service) call HQ 📞 050-3183-8835; for non-urgent issues direct to the inquiry form (Cast App → 問い合わせフォーム).
4. Always respond in English. Be concise. Number steps clearly.

KNOWLEDGE BASE:
{KNOWLEDGE}

TROUBLE FLOW:
{TROUBLE_FLOW}

QUESTION: {question}"""
    return api_call_with_retry(lambda: client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}]
    ).content[0].text)


def get_bot_response_b(question: str) -> str:
    """Part B: KNOWLEDGE + TROUBLE_FLOW のみ（新アーキテクチャ・Botと同一）"""
    return get_bot_response_a(question)


def _eval_sources_a() -> str:
    """Part A 評価者ソース: KNOWLEDGEとTROUBLE_FLOWのみ（Botと同じ）"""
    return f"=== KNOWLEDGE BASE ===\n{KNOWLEDGE}\n\n=== TROUBLE FLOW ===\n{TROUBLE_FLOW}"


def _eval_sources_b(question: str) -> str:
    """Part B 評価者ソース: KNOWLEDGE + TROUBLE_FLOW のみ（新アーキテクチャ）"""
    return _eval_sources_a()


def evaluate_response(question: str, response: str, category: str, sources: str) -> tuple[str, str]:
    if category in GTN_CATEGORIES:
        type_instruction = (
            "This is a life-admin question (banking, housing, SIM, etc.).\n"
            "CORRECT answer: direct the cast to GTN via their app.\n"
            "Answering with specific procedures = WRONG. HQ escalation = WRONG."
        )
        source_section = ""
    elif category in TYPE_B_CATEGORIES:
        type_instruction = (
            "This is a TYPE B general knowledge question (navigation, common sense, etc.).\n"
            "Evaluate whether the answer is practically correct and helpful.\n"
            "HQ escalation is NOT the expected answer for these questions."
        )
        source_section = ""
    else:
        type_instruction = (
            "This is a TYPE A business procedure question.\n"
            "The REFERENCE SOURCES below are the bot's ground truth.\n"
            "Penalize: (1) answers that invent procedures not in the sources,\n"
            "(2) unnecessary HQ escalation when the answer IS clearly in the sources."
        )
        source_section = f"\n=== REFERENCE SOURCES ===\n{sources}\n"

    prompt = f"""You are evaluating a support bot response for JollyCast cast members (CaSy, Japan).
CaSy support number 050-3183-8835 is correct and appropriate to include.

{type_instruction}
{source_section}
◎ = Accurate, specific, immediately actionable
○ = Mostly correct, minor gaps
△ = Partially correct, could mislead
✕ = Wrong or unable to answer

QUESTION: {question}

BOT RESPONSE:
{response}

Reply in this exact format (2 lines only):
RATING: [◎/○/△/✕]
COMMENT: [1-2 sentences in Japanese]"""
    result = api_call_with_retry(lambda: client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}]
    ).content[0].text.strip())
    rating, comment = "", ""
    for line in result.splitlines():
        if line.startswith("RATING:"):
            rating = line.replace("RATING:", "").strip()
        elif line.startswith("COMMENT:"):
            comment = line.replace("COMMENT:", "").strip()
    return rating, comment


# ── Excel書式 ────────────────────────────────────────────────
CAT_COLORS = {
    "ダブルブッキング": "E8F0FF", "サービス後・移動": "FFFAE4",
    "キャンセル詳細": "FFF0E0", "不衛生・退出判断": "FFE4CC",
    "自治体バウチャー": "CCF0FF", "QR・バウチャー": "E4FFFF",
    "料理サービス・返金": "FFF0F8", "お客様不在": "DDEEFF",
    "物損・事故": "FFE4E4", "HQエスカレーション": "F0FFE0",
    "アプリ操作・報告": "E8F4FF", "鍵・入室トラブル": "F0E4FF",
    "スケジュール変更": "FFE4F0",
    "GTN・送金": "E8FFE8", "GTN・銀行": "D4F5D4", "GTN・携帯": "C0EBC0",
}
RATING_COLORS = {"◎": "D4EDDA", "○": "D1ECF1", "△": "FFF3CD", "✕": "F8D7DA"}
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, color="2E4057"):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = PatternFill("solid", start_color=color, end_color=color)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 30


def apply_row_style(ws, row_idx, cat, rating, num_cols):
    cat_color = CAT_COLORS.get(cat, "FFFFFF")
    r_color = RATING_COLORS.get(rating, "FFFFFF")
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.fill = PatternFill("solid", start_color=(r_color if col == 5 else cat_color),
                             end_color=(r_color if col == 5 else cat_color))
        c.font = Font(name="Arial", size=9)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="top")
    ws.row_dimensions[row_idx].height = 80


def create_sheet(wb, name, questions, get_response_fn, get_eval_sources_fn, wait_sec):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    headers = ["#", "カテゴリ", "質問", "ボットの回答", "AI評価", "AI評価コメント", "担当者評価", "担当者コメント"]
    style_header(ws, headers)
    for col, w in zip(range(1, 9), [4, 22, 45, 55, 8, 38, 10, 30]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    ratings = []
    total = len(questions)
    for i, (cat, q) in enumerate(questions, 1):
        print(f"  [{i}/{total}] {cat}: {q[:50]}...", flush=True)
        if wait_sec > 0 and i > 1:
            print(f"  待機 {wait_sec}秒...", flush=True)
            time.sleep(wait_sec)
        response = get_response_fn(q)
        time.sleep(2)
        sources = get_eval_sources_fn(q)
        rating, comment = evaluate_response(q, response, cat, sources)
        time.sleep(2)
        print(f"  -> {rating}", flush=True)
        ratings.append((cat, rating))
        ws.append([i, cat, q, response, rating, comment, "", ""])
        apply_row_style(ws, ws.max_row, cat, rating, len(headers))
        wb.save(base / "jollycast_bot_test_results_v2.xlsx")

    # サマリー
    ws.append([])
    ws.append(["", "【サマリー】"])
    cat_r = defaultdict(list)
    for cat, r in ratings:
        cat_r[cat].append(r)
    for cat, rs in cat_r.items():
        cnt = Counter(rs)
        n = len(rs)
        good = cnt.get("◎", 0) + cnt.get("○", 0)
        ws.append(["", cat,
                   f"◎{cnt.get('◎',0)} ○{cnt.get('○',0)} △{cnt.get('△',0)} ✕{cnt.get('✕',0)}",
                   "", f"{round(good/n*100)}%"])
    all_cnt = Counter(r for _, r in ratings)
    good_all = all_cnt.get("◎", 0) + all_cnt.get("○", 0)
    ws.append(["", "【合計】",
               f"◎{all_cnt.get('◎',0)} ○{all_cnt.get('○',0)} △{all_cnt.get('△',0)} ✕{all_cnt.get('✕',0)}",
               "", f"{round(good_all/total*100)}%"])
    wb.save(base / "jollycast_bot_test_results_v2.xlsx")
    return ratings


def print_summary(name, ratings):
    all_cnt = Counter(r for _, r in ratings)
    total = len(ratings)
    good = all_cnt.get("◎", 0) + all_cnt.get("○", 0)
    print(f"\n=== {name} ===")
    cat_r = defaultdict(list)
    for cat, r in ratings:
        cat_r[cat].append(r)
    for cat, rs in cat_r.items():
        cnt = Counter(rs)
        n = len(rs)
        g = cnt.get("◎", 0) + cnt.get("○", 0)
        print(f"  {cat}: ◎{cnt.get('◎',0)} ○{cnt.get('○',0)} △{cnt.get('△',0)} ✕{cnt.get('✕',0)} → {round(g/n*100)}%")
    print(f"  全体: {round(good/total*100)}% ({good}/{total})")


def main():
    results_file = base / "jollycast_bot_test_results_v2.xlsx"
    wb = load_workbook(results_file)

    print("\n" + "="*50)
    print("Part A: 20問 KNOWLEDGE方式")
    print("="*50)
    ratings_a = create_sheet(wb, "Spot-A（KNOWLEDGE検証20問）",
                              QUESTIONS_A, get_bot_response_a,
                              lambda q: _eval_sources_a(), wait_sec=3)
    print_summary("Part A", ratings_a)

    print("\n" + "="*50)
    print("Part B: 5問 全Zendesk方式（150秒間隔）")
    print("="*50)
    ratings_b = create_sheet(wb, "Spot-B（Zendesk参照検証5問）",
                              QUESTIONS_B, get_bot_response_b,
                              _eval_sources_b, wait_sec=150)
    print_summary("Part B", ratings_b)

    print("\n" + "="*50)
    print("Part C: 3問 GTN案内検証")
    print("="*50)
    ratings_c = create_sheet(wb, "Spot-C（GTN案内検証3問）",
                              QUESTIONS_C, get_bot_response_a,
                              lambda q: "", wait_sec=3)
    print_summary("Part C", ratings_c)

    print(f"\nExcel保存完了: {results_file}")


if __name__ == "__main__":
    main()
